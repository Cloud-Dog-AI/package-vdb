# Copyright 2026 Cloud-Dog, Viewdeck Engineering Limited
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""In-memory workbook builders for the test matrix (requirements section 19.1).

These produce real ``.xlsx`` / ``.ods`` bytes so tests exercise the genuine
parsers without committing binary fixtures. They are importable by any service's
test suite (e.g. index-retriever) for shared, reproducible spreadsheet fixtures.
"""

from __future__ import annotations

import datetime as _dt
import io


def build_simple_xlsx() -> bytes:
    """Small simple workbook with a single formal table (section 19.1)."""
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["ID", "Customer", "Amount", "Date"])
    ws.append([1, "Acme Ltd", 1200.50, _dt.datetime(2026, 1, 10)])
    ws.append([2, "Globex", 980.00, _dt.datetime(2026, 1, 11)])
    ws.append([3, "Initech", 4300.75, _dt.datetime(2026, 1, 12)])
    table = Table(displayName="tbl_sales", ref="A1:D4")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)
    return _save_xlsx(wb)


def build_multisheet_formal_tables_xlsx() -> bytes:
    """Workbook with multiple sheets and formal tables (section 19.1)."""
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Revenue"
    ws1.append(["Region", "Q1", "Q2"])
    ws1.append(["North", 100, 120])
    ws1.append(["South", 90, 95])
    ws1.add_table(Table(displayName="tbl_revenue", ref="A1:C3"))

    ws2 = wb.create_sheet("Costs")
    ws2.append(["Category", "Amount"])
    ws2.append(["Travel", 4000])
    ws2.append(["Software", 12000])
    ws2.add_table(Table(displayName="tbl_costs", ref="A1:B3"))
    return _save_xlsx(wb)


def build_inferred_only_xlsx() -> bytes:
    """Workbook whose data has NO formal tables, plus a note + decorative cell."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Raw"
    # Inferred dataset block at A1.
    ws.append(["Product", "Units", "Price"])
    ws.append(["Widget", 10, 2.5])
    ws.append(["Gadget", 5, 9.99])
    ws.append(["Sprocket", 22, 1.25])
    # blank separator rows
    ws.append([])
    ws.append([])
    # A free-form note block lower down.
    ws["A7"] = "Notes: these figures are provisional and subject to revision pending audit review."
    ws["A8"] = "Contact the finance team for the authoritative numbers before circulation."
    return _save_xlsx(wb)


def build_hidden_sheet_xlsx() -> bytes:
    """Workbook with a visible and a hidden sheet (section 19.1)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Public"
    ws.append(["Key", "Value"])
    ws.append(["alpha", 1])
    hidden = wb.create_sheet("Secret")
    hidden.append(["Token", "Owner"])
    hidden.append(["xyz", "ops"])
    hidden.sheet_state = "hidden"
    return _save_xlsx(wb)


def build_formula_named_range_xlsx() -> bytes:
    """Workbook with formulas and a workbook-scoped named range (section 19.1)."""
    from openpyxl import Workbook
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()
    ws = wb.active
    ws.title = "Calc"
    ws.append(["Item", "Value"])
    ws.append(["a", 10])
    ws.append(["b", 20])
    ws.append(["c", 30])
    ws["B5"] = "=SUM(B2:B4)"
    ws["A5"] = "Total"
    wb.defined_names.add(DefinedName("ValueColumn", attr_text="Calc!$B$2:$B$4"))
    return _save_xlsx(wb)


def build_multilingual_xlsx() -> bytes:
    """Multilingual workbook: CJK, RTL and accented text (section 5.19, 19.1)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "国际"
    ws.append(["名称", "Montant", "تاريخ"])
    ws.append(["商品A", 1234.56, "2026-01-01"])
    ws.append(["café", 9876.54, "2026-02-02"])
    ws.append(["מוצר", 4321.00, "2026-03-03"])
    return _save_xlsx(wb)


def build_mixed_merged_xlsx() -> bytes:
    """Workbook with a merged title row above a mixed-type table (section 19.1)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Quarterly Performance Report"
    ws.merge_cells("A1:D1")
    ws.append([])  # row 2 spacer pushed by explicit row 1; build header at row 3
    ws["A3"] = "Name"
    ws["B3"] = "Active"
    ws["C3"] = "Score"
    ws["D3"] = "Joined"
    ws.append(["Alice", True, 91.5, _dt.datetime(2025, 6, 1)])
    ws.append(["Bob", False, 73, _dt.datetime(2025, 7, 15)])
    return _save_xlsx(wb)


def build_report_grid_xlsx() -> bytes:
    """Workbook with a merged title above a label-column numeric matrix (cross-tab)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Grid"
    ws["A1"] = "Sales by Region (USD thousands)"
    ws.merge_cells("A1:D1")
    ws["A2"] = "Region"
    ws["B2"] = "Q1"
    ws["C2"] = "Q2"
    ws["D2"] = "Q3"
    ws.append(["North", 100, 120, 130])
    ws.append(["South", 90, 95, 99])
    ws.append(["East", 70, 80, 85])
    return _save_xlsx(wb)


def build_large_xlsx(rows: int = 500) -> bytes:
    """Large workbook to exercise row-batch indexing (section 13.2, 19.1)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Big"
    ws.append(["RowId", "Label", "Measure"])
    for i in range(1, rows + 1):
        ws.append([i, f"label-{i}", i * 1.5])
    return _save_xlsx(wb)


def build_malformed_bytes() -> bytes:
    """Bytes that are not a valid workbook (section 19.1 partial recovery)."""
    return b"PK\x03\x04 this is not really a valid spreadsheet payload at all"


def build_simple_ods() -> bytes:
    """Small ODS workbook with a float column, a date column and a named range."""
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import NamedExpressions, NamedRange, Table, TableCell, TableRow
    from odf.text import P

    doc = OpenDocumentSpreadsheet()

    def _row(values: list[object]) -> TableRow:
        tr = TableRow()
        for value in values:
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell = TableCell(valuetype="float", value=value)
            elif isinstance(value, _dt.date):
                cell = TableCell(valuetype="date", datevalue=value.isoformat())
            else:
                cell = TableCell(valuetype="string")
            cell.addElement(P(text="" if value is None else str(value)))
            tr.addElement(cell)
        return tr

    sheet = Table(name="Budget")
    sheet.addElement(_row(["Item", "Amount", "When"]))
    sheet.addElement(_row(["rent", 1000.0, _dt.date(2026, 1, 1)]))
    sheet.addElement(_row(["power", 250.5, _dt.date(2026, 1, 2)]))
    sheet.addElement(_row(["water", 75.25, _dt.date(2026, 1, 3)]))
    doc.spreadsheet.addElement(sheet)

    sheet2 = Table(name="Meta")
    sheet2.addElement(_row(["note", "second sheet"]))
    doc.spreadsheet.addElement(sheet2)

    named = NamedExpressions()
    named.addElement(NamedRange(name="AmountCol", cellrangeaddress="Budget.B2:Budget.B4"))
    doc.spreadsheet.addElement(named)

    buffer = io.BytesIO()
    doc.write(buffer)
    return buffer.getvalue()


def _save_xlsx(workbook: object) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)  # type: ignore[attr-defined]
    return buffer.getvalue()
