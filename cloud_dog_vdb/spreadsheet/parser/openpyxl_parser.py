# Copyright 2026 Cloud-Dog, Viewdeck Engineering Limited
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""OOXML (``.xlsx`` / ``.xlsm``) workbook parser built on openpyxl (section 11).

The parser is strictly passive: it loads the workbook with VBA, external links
and macro evaluation disabled, and never executes any active workbook content
(section 17.1). Two loads are performed — one with ``data_only=False`` to keep
formula text, and one with ``data_only=True`` to recover the cached display
values for formula cells.
"""

from __future__ import annotations

import io
import zipfile
from typing import Any

import openpyxl
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.formula import ArrayFormula

from cloud_dog_vdb.spreadsheet.config import SpreadsheetConfig
from cloud_dog_vdb.spreadsheet.parser.base import (
    RawFormula,
    RawNamedRange,
    RawPivot,
    RawSheet,
    RawTable,
    RawWorkbook,
    WorkbookParser,
)

#: Marker entry whose presence in the OOXML zip indicates an embedded macro project.
_VBA_PROJECT_PATH = "xl/vbaProject.bin"


class OpenpyxlParser(WorkbookParser):
    """Parse ``.xlsx`` / ``.xlsm`` bytes into a :class:`RawWorkbook` without running macros."""

    supported_formats = ("xlsx", "xlsm")

    def parse(self, data: bytes, *, file_name: str, config: SpreadsheetConfig) -> RawWorkbook:
        """Parse workbook ``data`` into a :class:`RawWorkbook` (no macro execution)."""
        is_xlsm = file_name.lower().endswith(".xlsm")
        workbook = RawWorkbook(
            file_name=file_name,
            file_format="xlsm" if is_xlsm else "xlsx",
            has_macros=is_xlsm or _has_vba_project(data),
        )

        # Primary load keeps formula text and never evaluates macros / external links.
        wb = openpyxl.load_workbook(
            io.BytesIO(data),
            data_only=False,
            read_only=False,
            keep_vba=False,
            keep_links=False,
        )

        # Secondary load surfaces cached display values for formula cells. It is
        # best-effort: a failure here simply leaves formula display values empty.
        wb_values = None
        try:
            wb_values = openpyxl.load_workbook(
                io.BytesIO(data),
                data_only=True,
                read_only=False,
                keep_vba=False,
                keep_links=False,
            )
        except Exception:  # noqa: BLE001 - cached values are optional
            wb_values = None

        workbook.properties = _extract_properties(wb, file_format=workbook.file_format)

        for index, ws in enumerate(wb.worksheets):
            try:
                values_ws = _matching_values_sheet(wb_values, ws.title)
                sheet = _parse_sheet(ws, index, values_ws, config, workbook)
                workbook.sheets.append(sheet)
            except Exception as err:  # noqa: BLE001 - graceful per-sheet degradation
                workbook.warnings.append(f"sheet {ws.title}: {err}")
                workbook.parse_status = "partial"

        workbook.named_ranges = _extract_workbook_named_ranges(wb)

        return workbook


def _has_vba_project(data: bytes) -> bool:
    """Return ``True`` if the OOXML zip embeds a VBA macro project."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            return _VBA_PROJECT_PATH in zf.namelist()
    except Exception:  # noqa: BLE001 - non-zip / corrupt input is simply "no macros"
        return False


def _extract_properties(wb: Any, *, file_format: str) -> dict[str, Any]:
    """Capture available workbook core properties as plain scalars/ISO strings."""
    props: dict[str, Any] = {"file_format": file_format}
    source = getattr(wb, "properties", None)
    for attr in ("title", "subject", "creator", "keywords", "description", "category"):
        try:
            value = getattr(source, attr, None)
        except Exception:  # noqa: BLE001 - defensive per-field guard
            value = None
        if value is not None:
            props[attr] = value
    for attr in ("created", "modified"):
        try:
            value = getattr(source, attr, None)
        except Exception:  # noqa: BLE001 - defensive per-field guard
            value = None
        if value is not None:
            try:
                props[attr] = value.isoformat()
            except Exception:  # noqa: BLE001 - fall back to str() for odd types
                props[attr] = str(value)
    return props


def _matching_values_sheet(wb_values: Any, title: str) -> Any:
    """Return the ``data_only`` worksheet matching ``title`` (or ``None``)."""
    if wb_values is None:
        return None
    try:
        if title in wb_values.sheetnames:
            return wb_values[title]
    except Exception:  # noqa: BLE001 - best-effort lookup
        return None
    return None


def _parse_sheet(
    ws: Any,
    index: int,
    values_ws: Any,
    config: SpreadsheetConfig,
    workbook: RawWorkbook,
) -> RawSheet:
    """Build a :class:`RawSheet` for ``ws`` (A1-origin dense grid + metadata)."""
    sheet = RawSheet(
        name=ws.title,
        index=index,
        visibility=ws.sheet_state,
        used_range=ws.dimensions or "",
        merged_cells=[str(r) for r in ws.merged_cells.ranges],
        hidden_rows=_hidden_indexes(ws.row_dimensions),
        hidden_columns=_hidden_column_indexes(ws.column_dimensions),
        frozen_panes=ws.freeze_panes or "",
    )

    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # Scan cap: bound the number of rows so row*col stays under the configured cap.
    cap = config.max_cells_per_sheet_scan
    read_rows = max_row
    if max_row and max_col and max_row * max_col > cap:
        read_rows = max(0, cap // max_col) if max_col else 0
        workbook.warnings.append(
            f"sheet {ws.title}: scan capped at {read_rows} of {max_row} rows "
            f"(cells {max_row * max_col} exceed cap {cap})"
        )
        workbook.parse_status = "partial"

    if read_rows and max_col:
        sheet.cells = [
            list(row)
            for row in ws.iter_rows(
                min_row=1,
                min_col=1,
                max_row=read_rows,
                max_col=max_col,
                values_only=True,
            )
        ]

    sheet.formal_tables = _extract_tables(ws)
    sheet.formulas = _extract_formulas(ws, values_ws, read_rows, max_col)
    sheet.pivots = _extract_pivots(ws)
    return sheet


def _extract_pivots(ws: Any) -> list[RawPivot]:
    """Best-effort pivot extraction from openpyxl's parsed pivot tables (section 5.9)."""
    pivots: list[RawPivot] = []
    for definition in getattr(ws, "_pivots", None) or []:
        try:
            pivots.append(_pivot_from_definition(definition))
        except Exception:  # noqa: BLE001 - pivot detail is best-effort
            continue
    return pivots


def _pivot_from_definition(definition: Any) -> RawPivot:
    """Map an openpyxl pivot ``TableDefinition`` to a :class:`RawPivot`."""
    cache = getattr(definition, "cache", None)
    cache_fields: list[str] = []
    source_ref = ""
    if cache is not None:
        for field in getattr(cache, "cacheFields", None) or []:
            cache_fields.append(getattr(field, "name", "") or "")
        cache_source = getattr(cache, "cacheSource", None)
        worksheet_source = getattr(cache_source, "worksheetSource", None) if cache_source else None
        if worksheet_source is not None:
            source_ref = getattr(worksheet_source, "ref", "") or getattr(worksheet_source, "name", "") or ""

    def _name(index: Any) -> str | None:
        if index is None or not isinstance(index, int) or index < 0:
            return None
        if index < len(cache_fields) and cache_fields[index]:
            return cache_fields[index]
        return f"field_{index}"

    row_fields = [n for n in (_name(getattr(f, "x", None)) for f in getattr(definition, "rowFields", None) or []) if n]
    col_fields = [n for n in (_name(getattr(f, "x", None)) for f in getattr(definition, "colFields", None) or []) if n]
    page_fields = [
        n for n in (_name(getattr(f, "fld", None)) for f in getattr(definition, "pageFields", None) or []) if n
    ]
    value_fields: list[str] = []
    for data_field in getattr(definition, "dataFields", None) or []:
        name = getattr(data_field, "name", None) or _name(getattr(data_field, "fld", None))
        if name:
            value_fields.append(name)

    location = getattr(definition, "location", None)
    range_ref = (getattr(location, "ref", "") or "") if location is not None else ""
    completeness = "complete" if (row_fields or value_fields) else "partial"
    return RawPivot(
        name=getattr(definition, "name", "") or "pivot",
        source_ref=source_ref,
        row_fields=row_fields,
        column_fields=col_fields,
        value_fields=value_fields,
        filter_fields=page_fields,
        range_ref=range_ref,
        extraction_completeness=completeness,
    )


def _hidden_indexes(dimensions: Any) -> list[int]:
    """Return sorted zero-based row indexes for hidden row dimensions."""
    hidden: list[int] = []
    for key, dim in dimensions.items():
        if getattr(dim, "hidden", False):
            try:
                hidden.append(int(key) - 1)
            except (TypeError, ValueError):
                continue
    return sorted(set(hidden))


def _hidden_column_indexes(dimensions: Any) -> list[int]:
    """Return sorted zero-based column indexes for hidden column dimensions."""
    from openpyxl.utils.cell import column_index_from_string

    hidden: list[int] = []
    for key, dim in dimensions.items():
        if not getattr(dim, "hidden", False):
            continue
        # A column dimension may cover a span (min..max); expand it.
        col_min = getattr(dim, "min", None)
        col_max = getattr(dim, "max", None)
        if col_min and col_max:
            hidden.extend(range(int(col_min) - 1, int(col_max)))
            continue
        try:
            hidden.append(column_index_from_string(str(key)) - 1)
        except (TypeError, ValueError):
            continue
    return sorted(set(hidden))


def _extract_tables(ws: Any) -> list[RawTable]:
    """Extract formal tables declared on the worksheet (section 5.4)."""
    tables: list[RawTable] = []
    table_map = getattr(ws, "tables", {}) or {}
    for tbl in table_map.values():
        ref = getattr(tbl, "ref", "") or ""
        header_row_index, column_labels = _table_header(ws, ref)
        style_name = ""
        try:
            style_info = getattr(tbl, "tableStyleInfo", None)
            if style_info is not None:
                style_name = getattr(style_info, "name", "") or ""
        except Exception:  # noqa: BLE001 - style metadata is optional
            style_name = ""
        totals_present = bool(getattr(tbl, "totalsRowCount", 0) or getattr(tbl, "totalsRowShown", False))
        tables.append(
            RawTable(
                name=tbl.name,
                display_name=getattr(tbl, "displayName", "") or tbl.name,
                range_ref=ref,
                style_name=style_name,
                header_row_index=header_row_index,
                totals_row_present=totals_present,
                column_labels=column_labels,
            )
        )
    return tables


def _table_header(ws: Any, ref: str) -> tuple[int, list[str]]:
    """Return ``(zero_based_header_row, header_labels)`` for a table ``ref``."""
    if not ref:
        return 0, []
    try:
        min_col, min_row, max_col, _max_row = range_boundaries(ref)
    except Exception:  # noqa: BLE001 - malformed ref → no header info
        return 0, []
    labels: list[str] = []
    for col in range(min_col, max_col + 1):
        value = ws.cell(row=min_row, column=col).value
        labels.append("" if value is None else str(value))
    return min_row - 1, labels


def _extract_formulas(ws: Any, values_ws: Any, read_rows: int, max_col: int) -> list[RawFormula]:
    """Capture formula cells (text only, never evaluated) with cached display values."""
    formulas: list[RawFormula] = []
    if not read_rows or not max_col:
        return formulas
    for row in ws.iter_rows(min_row=1, min_col=1, max_row=read_rows, max_col=max_col):
        for cell in row:
            formula_text = _formula_text(cell)
            if formula_text is None:
                continue
            display_value = ""
            if values_ws is not None:
                try:
                    cached = values_ws.cell(row=cell.row, column=cell.column).value
                    if cached is not None:
                        display_value = str(cached)
                except Exception:  # noqa: BLE001 - cached lookup is best-effort
                    display_value = ""
            formulas.append(
                RawFormula(
                    cell_ref=cell.coordinate,
                    formula_text=formula_text,
                    display_value=display_value,
                )
            )
    return formulas


def _formula_text(cell: Any) -> str | None:
    """Return the formula string for ``cell`` or ``None`` if it is not a formula."""
    value = cell.value
    if isinstance(value, ArrayFormula):
        return getattr(value, "text", "") or ""
    if getattr(cell, "data_type", None) == "f":
        return value if isinstance(value, str) else (str(value) if value is not None else "")
    if isinstance(value, str) and value.startswith("="):
        return value
    return None


def _extract_workbook_named_ranges(wb: Any) -> list[RawNamedRange]:
    """Collect workbook-scope and per-sheet named ranges, skipping reserved names."""
    ranges: list[RawNamedRange] = []
    ranges.extend(_named_ranges_from(wb.defined_names, scope="workbook"))
    for ws in wb.worksheets:
        sheet_names = getattr(ws, "defined_names", None)
        if sheet_names:
            ranges.extend(_named_ranges_from(sheet_names, scope=ws.title))
    return ranges


def _named_ranges_from(defined_names: Any, *, scope: str) -> list[RawNamedRange]:
    """Convert a dict-like ``DefinedNameDict`` into :class:`RawNamedRange` entries."""
    result: list[RawNamedRange] = []
    try:
        items = list(defined_names.items())
    except Exception:  # noqa: BLE001 - tolerate unexpected container shapes
        items = []
    for name, dn in items:
        dn_name = getattr(dn, "name", None) or name
        if dn_name.startswith("_xlnm") or getattr(dn, "is_reserved", False):
            continue
        refers_to = getattr(dn, "value", None) or getattr(dn, "attr_text", None) or ""
        result.append(RawNamedRange(name=dn_name, refers_to=refers_to, scope=scope))
    return result
